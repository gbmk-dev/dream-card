#-*- coding:utf-8 -*-
import os
from openpyxl import load_workbook
from pymongo import MongoClient           # pymongo를 임포트 하기(패키지 인스톨 먼저 해야겠죠?)
from openpyxl.workbook import Workbook
import xlrd

client = MongoClient('localhost', 27017)  # mongoDB는 27017 포트로 돌아갑니다.
db = client.dreamCard                    # 'dreamCard'라는 이름의 db를 만듭니다.


def convert_xls_to_xlsx(filename):
    # first open using xlrd
    book = xlrd.open_workbook(filename)
    xlsx_filename = filename[:filename.rfind('.')]
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.active

    for row in range(0, nrows):
        for col in range(0, ncols):
            sheet1.cell(row=row + 1, column=col + 1).value = sheet.cell_value(row, col)

    book1.save(filename='{}.xlsx'.format(xlsx_filename))
    return book1


convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/1.종로구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/2.중구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/3.용산구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/4.성동구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/5.광진구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/6.동대문구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/7.중랑구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/8.성북구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/9.강북구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/10.도봉구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/11.노원구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/12.은평구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/13.서대문구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/14.마포구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/15.양천구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/16.강서구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/17.구로구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/18.금천구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/19.영등포구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/20.동작구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/21.관악구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/22.서초구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/23.강남구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/24.송파구.xls')
convert_xls_to_xlsx('/Users/giubinkang/Desktop/projects/dream_card_project/db/25.강동구.xls')

cwd = os.getcwd()  # Get the current working directory (cwd)
files_collection = os.listdir(cwd)  # Get all the files in that directory
xlsx_files = [file for file in files_collection if file.endswith('.xlsx')]

num = 0
for county_file_name in xlsx_files:
    file_name = cwd + '/' + county_file_name  # 맥, 윈도우 계속 수정해야한다.
    work_book = load_workbook(file_name, data_only=True)  # 파일의 제목을 읽어오는 작업
    work_sheet = work_book['Sheet']  # 엑셀 파일을 열었을때 안에 있는 정보들을 읽어와라
    print(work_book)

    for col in work_sheet.iter_rows(min_row=4, max_row=100000):

        if col[1].value is not None:
            restaurant_name = col[1].value
        else:
            continue
        if col[2].value is not None:
            restaurant_type = col[2].value
        else:
            continue
        if col[3].value is not None:
            restaurant_phoneNumber = col[3].value
        else:
            continue
        if col[4].value is not None:
            restaurant_address = col[4].value
        else:
            continue
        if work_sheet is not None:
            county_name = county_file_name.split('.')[1]
        else:
            continue


        # restaurant_name = col[1].value if col[1].value is not None else '0'
        # restaurant_type = col[2].value if col[2].value is not None else '0'
        # restaurant_phoneNumber = col[3].value if col[3].value is not None else '0'
        # restaurant_address = col[4].value if col[4].value is not None else '0'

        restaurants_infos = {
            'restaurant_name': restaurant_name,
            'restaurant_type': restaurant_type,
            'restaurant_phoneNumber': restaurant_phoneNumber,
            'restaurant_address': restaurant_address,
            'county_name': county_name
        }
        num += 1
        db.county_files.insert_one(restaurants_infos) #정보들 insert 하기
        #print(num, restaurants_infos)
        #print(col[1].value, col[2].value, col[3].value, col[4].value)
        print(num, col[1].value,col[4].value, county_name)
